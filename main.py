import random
import os
import config as c
from secret import run_algo
from openpyxl import Workbook
from openpyxl.styles import *

wb = Workbook()
sheet = wb.active


for i in range(c.ZONE_LENGTH * c.ZONE_NUMBER):
    c.CONVEYOR.append(0)
    c.EMPTY_CONVEYOR.append(0)
CONVEYOR_LENGTH = len(c.CONVEYOR)

for i in range(c.ZONE_NUMBER):
    c.ZONE_WEIGHTS.append(0)
    c.ZONE_RATING.append("L")
    c.SPEED.append(1)
    c.SPEED_RATING.append("-")

#data collection vars
WEIGHT_SPREAD = [0, 0, 0, 0]
TOTAL_FRAMES = 0

if __name__ == '__main__':
    print("Running Simulation")
    for run_count in range(1,c.RUN_TIMES):
        if (run_count > 32) & (c.CONVEYOR == c.EMPTY_CONVEYOR):
            break
        OUTPUT = 0
        if (c.SPEED[7] == 0.5):
            if (c.CONVEYOR[CONVEYOR_LENGTH - 1] > 0):
                OUTPUT += 0.25
                c.CONVEYOR[CONVEYOR_LENGTH - 1] -= 0.25
        else:
            for i in range(1, c.SPEED[7] + 1):
                OUTPUT += c.CONVEYOR[CONVEYOR_LENGTH - i]
                c.CONVEYOR[CONVEYOR_LENGTH - i] = 0
        if OUTPUT == 0:
            WEIGHT_SPREAD[0] += 1
        if OUTPUT == 0.25:
            WEIGHT_SPREAD[1] += 1
        if OUTPUT == 0.50:
            WEIGHT_SPREAD[2] += 1
        if OUTPUT == 0.75:
            WEIGHT_SPREAD[3] += 1


        #move conveyor
        for a in range(CONVEYOR_LENGTH - 1, 0, -1):
            if (run_count % 2 == 0) & (a - 1 >= 0) & (c.SPEED[int((a - 1) / c.ZONE_LENGTH)] == 0.5):
                if (c.CONVEYOR[a - 1] > 0) & (c.CONVEYOR[a] < .25):
                    c.CONVEYOR[a] += 0.25
                    c.CONVEYOR[a - 1] -= 0.25
            for i in range(1,4):
                if (a - i >= 0) & (i == c.SPEED[int((a - i) / c.ZONE_LENGTH)]) & \
                        (c.CONVEYOR[a - i] > 0) & (c.CONVEYOR[a] < .25):
                    c.CONVEYOR[a] += 0.25
                    c.CONVEYOR[a - i] -= 0.25
        TOTAL_FRAMES += 1

        # select input types
        if c.INPUT == "random":
            CONVEYOR_ADD_OPTIONS = [0, .25,]
            # sets number of inputs
            if run_count < c.INPUT_NUM:
                c.CONVEYOR[0] += random.choices(CONVEYOR_ADD_OPTIONS, weights=(c.PROBS[0], c.PROBS[1]))[0]
        if c.INPUT == "manual":
            if run_count < len(c.CONVEYOR_ADD) - 2:
                c.CONVEYOR[0] = c.CONVEYOR_ADD[run_count]
        if c.CONVEYOR[0] > 0.75:
            print("OVERLOAD")

        #calculate zone weights
        c.ZONE_WEIGHTS = []
        for i in range(c.ZONE_NUMBER):
            c.ZONE_WEIGHTS.append(0)
        for zone in range(CONVEYOR_LENGTH):
            c.ZONE_WEIGHTS[int(zone/c.ZONE_LENGTH)] += c.CONVEYOR[zone]

        #assign rating to zone depending on weight
        for a in range(len(c.ZONE_WEIGHTS)):
            if c.ZONE_WEIGHTS[a] < 0.5:
                c.ZONE_RATING[a] = "L"
            elif c.ZONE_WEIGHTS[a] < 1.5:
                c.ZONE_RATING[a] = "N"
            else:
                c.ZONE_RATING[a] = "H"

        #assign c.speed to zone depending on rating
        c.SPEED = []
        c.SPEED_RATING = []
        for i in range(c.ZONE_NUMBER):
            c.SPEED.append(1)
            c.SPEED_RATING.append("-")
        SUB = " "
        if c.ALGO_ON:
            for x in range(1, len(c.ZONE_WEIGHTS) - 1):
                run_algo(x)
                SUB += "; "
                #c.SPEED = [0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5,]
                c.SPEED_RATING[0] = "-"
                c.SPEED_RATING[len(c.SPEED_RATING) - 1] = c.SPEED_RATING[len(c.SPEED_RATING) - 2]
                c.SPEED[0] = 1
                c.SPEED[len(c.SPEED) - 1] = c.SPEED[len(c.SPEED) - 2]

        #write to spreadsheet
        if c.SPREADSHEET:
            COLS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
                    "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM",
                    "AN", "AO", "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BD",
                    "BE", "BF", "BG", "BH", "BI", "BJ", "BK", "BL", "BM", "BN", "BO", "BP", "BQ", "BR", "BS", "BT", "BU",
                    "BV", "BW", "BX", "BY", "BZ", "CA", "CB", "CC", "CD", "CE", "CF", "CG", "CH", "CI", "CJ", "CK", "CL",
                    "CM", "CN", "CO", "CP", "CQ", "CR", "CS", "CT", "CU", "CV", "CW", "CX", "CY", "CZ", "DA", "DB", "DC",
                    "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL", "DM", "DN", "DO", "DP", "DQ", "DR", "DS", "DT",
                    "DU", "DV", "DW", "DX", "DY", "DZ", "EA", "EB", "EC", "ED", "EE", "EF", "EG", "EH", "EI", "EJ", "EK",
                    "EL", "EM", "EN", "EO", "EP", "EQ", "ER", "ES", "ET", "EU", "EV", "EW", "EX", "EY", "EZ", ]
            for z in range(len(c.CONVEYOR)):
                sheet[str(COLS[z]) + str(run_count * 4 + 1)] = str(c.CONVEYOR[z])
            for z in range(0, len(c.CONVEYOR) - 2, 3):
                if c.CONVEYOR[z] + c.CONVEYOR[z+1] + c.CONVEYOR[z+2] == 0.25:
                    sheet[str(COLS[z]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="E5E5E5",
                                                                                    fill_type = "solid")
                    sheet[str(COLS[z+1]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="E5E5E5",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+2]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="E5E5E5",
                                                                                    fill_type="solid")
                elif c.CONVEYOR[z] + c.CONVEYOR[z+1] + c.CONVEYOR[z+2] == 0.50:
                    sheet[str(COLS[z]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="CECECE",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+1]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="CECECE",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+2]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="CECECE",
                                                                                    fill_type="solid")
                elif c.CONVEYOR[z] + c.CONVEYOR[z+1] + c.CONVEYOR[z+2] == 0.75:
                    sheet[str(COLS[z]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="A6A6A6",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+1]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="A6A6A6",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+2]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="A6A6A6",
                                                                                    fill_type="solid")
                elif c.CONVEYOR[z] + c.CONVEYOR[z+1] + c.CONVEYOR[z+2] > 0.75:
                    sheet[str(COLS[z]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="FFB1B1",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+1]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="A6A6A6",
                                                                                    fill_type="solid")
                    sheet[str(COLS[z+2]) + str(run_count * 4 + 1)].fill = PatternFill(start_color="A6A6A6",
                                                                                    fill_type="solid")
            for z in range(3, len(c.CONVEYOR), c.ZONE_LENGTH):
                sheet[str(COLS[z - 3]) + str(run_count * 4)] = int(z / c.ZONE_LENGTH)
                sheet[str(COLS[z]) + str(run_count * 4 + 2)] = c.SPEED_RATING[int(z / c.ZONE_LENGTH)]
                sheet[str(COLS[z - 1]) + str(run_count * 4 + 2)] = c.ZONE_RATING[int(z / c.ZONE_LENGTH)]
                sheet[str(COLS[z - 2]) + str(run_count * 4 + 2)] = str(c.ZONE_WEIGHTS[int(z / c.ZONE_LENGTH)])
            sheet[str(COLS[CONVEYOR_LENGTH + 1]) + str(run_count * 4 + 1)] = str(OUTPUT)
            sheet[str(COLS[CONVEYOR_LENGTH + 2]) + str(run_count * 4 + 1)] = str(c.ZONE_RATING)
            sheet[str(COLS[CONVEYOR_LENGTH + 3]) + str(run_count * 4 + 1)] = SUB

    print("Weight Spread")
    WEIGHT_SPREAD = [WEIGHT_SPREAD[0] / run_count, WEIGHT_SPREAD[1] / run_count, WEIGHT_SPREAD[2] / run_count,
                     WEIGHT_SPREAD[3] / run_count, ]
    print(WEIGHT_SPREAD)
    print("Total Frames")
    print(TOTAL_FRAMES)
    if c.SPREADSHEET:
        sheet[(COLS[0]) + str(run_count * 4)] = "Weight Spread"
        sheet[(COLS[0]) + str(run_count * 4+ 1)] = str(WEIGHT_SPREAD)
        sheet[(COLS[0]) + str(run_count * 4+ 2)] = "Total Frames"
        sheet[(COLS[0]) + str(run_count * 4+ 3)] = TOTAL_FRAMES


    if c.SPREADSHEET:
        # styles
        if c.STYLE:
            print("Styling Spreadsheet")
            for i in range(1, run_count * 4):
                for l in COLS:
                    sheet.column_dimensions[l].width = 4
                    sheet.row_dimensions[i].height = 28
                    sheet[l + str(i)].alignment = Alignment(horizontal="center")
            for i in range(5, run_count * 4, 4):
                for z in range(CONVEYOR_LENGTH):
                    if z % c.ZONE_LENGTH == 0:
                        sheet[str(COLS[z]) + str(i)].border = Border(top=Side(border_style="medium"),
                                                                             bottom=Side(border_style="medium"),
                                                                             left=Side(border_style="medium"))
                    elif z % (c.ZONE_LENGTH / 4) == 0:
                        sheet[str(COLS[z]) + str(i)].border = Border(top=Side(border_style="medium"),
                                                                     bottom=Side(border_style="medium"),
                                                                     left=Side(border_style="thin"))
                    else:
                        sheet[str(COLS[z]) + str(i)].border = Border(top=Side(border_style="medium"),
                                                                             bottom=Side(border_style="medium"))
                print(i)

        print("Saving Spreadsheet")
        wb.save(filename="./scenarios/" + c.FILE + ".xlsx")
        print("Complete")
        os.system('say "Complete"')